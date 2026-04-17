#!/usr/bin/env python3
"""
json2tm.py — Convert JSON translation files (EN / DE / RU) to TMX and XLSX.

Field naming conventions understood
------------------------------------
  snake_case  :  label_en  →  label_de / label_ru
  camelCase   :  labelEn   →  labelDe  / labelRu

Cross-file matching
-------------------
The three JSON files must share the same structure.  For every *_en / *En
field found in the EN file the script looks for the peer *_de / *De field
at the SAME dict level inside the DE file (and likewise for RU).

If the DE / RU files carry the same field names as EN (e.g. all three files
use label_en but with different text), pass --same-keys and the value of
each *_en field is read from the corresponding file.

Usage — single files
---------------------
    python json2tm.py --en translations_en.json \\
                      --de translations_de.json \\
                      --ru translations_ru.json \\
                      [--out output/] [--same-keys] [--no-tmx] [--no-xlsx]

Usage — directories (recursive)
---------------------------------
    python json2tm.py --en path/to/en/ \\
                      --de path/to/de/ \\
                      --ru path/to/ru/ \\
                      [--out output/]

    Every *.json file found (recursively) under the EN directory is paired
    with the file at the same relative path inside the DE and RU directories.
    All triplets are processed in one run; segments are deduplicated globally.

Dependencies
------------
    pip install tqdm openpyxl
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, NamedTuple
import xml.etree.ElementTree as ET
from xml.dom import minidom

# ── Optional dependencies ─────────────────────────────────────────────────────

try:
    from tqdm import tqdm as _tqdm_cls
    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False

    class _tqdm_cls:  # type: ignore[no-redef]
        """Minimal no-op stand-in when tqdm is not installed."""
        def __init__(self, iterable=None, **kw):
            self._it = iterable
            total = kw.get("total")
            desc = kw.get("desc", "")
            if total:
                print(f"{desc}: processing {total} items …", file=sys.stderr)
            else:
                print(f"{desc} …", file=sys.stderr)

        def __enter__(self): return self
        def __exit__(self, *a): pass
        def update(self, n=1): pass
        def set_postfix_str(self, s, **kw): pass
        def __iter__(self): return iter(self._it or [])

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ── Constants ─────────────────────────────────────────────────────────────────

TOOL_NAME    = "json2tm"
TOOL_VERSION = "1.0"

UUID_RE = re.compile(
    r"^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$",
    re.IGNORECASE,
)
UUID_KEY_RE = re.compile(r"(?i)(uuid|_id)$")

# Language descriptors: (snake suffix, camelCase suffix, BCP-47 tag)
LANGS = {
    "en": ("_en", "En", "en"),
    "de": ("_de", "De", "de"),
    "ru": ("_ru", "Ru", "ru"),
}

# ── Text helpers ──────────────────────────────────────────────────────────────

def _clean_text(s: str) -> str:
    """Strip whitespace and leading BOM characters (U+FEFF) from a text value."""
    return s.strip().lstrip('\ufeff')


# ── QA patterns ───────────────────────────────────────────────────────────────

# Interpolation placeholders: {var}, {{var}}, %s, %d, %(name)s, etc.
_PH_RE       = re.compile(r'%[sdif%]|%\(\w+\)[sdif]|\{\{?[\w. ]+\}?\}')
_NUM_RE      = re.compile(r'\d+')
_TAG_RE      = re.compile(r'<[a-zA-Z/][^>]*>')
_CYRILLIC_RE = re.compile(r'[\u0400-\u04FF]')
_LATIN_LC_RE = re.compile(r'[a-z]')

# For TAG_MISMATCH: normalise tag attributes away so that <d=Holotopy> and
# <d=Голотопия> both compare as <d>, avoiding false positives on translated
# section labels.  Captures the tag name only (with optional leading /).
_TAG_NAME_RE = re.compile(r'^<(/?[a-zA-Z]+)')


def _strip_tag_attrs(tag: str) -> str:
    """Normalise a tag to its name: ``<d=Holotopy>`` → ``<d>``, ``<color=#6F6F6F>`` → ``<color>``."""
    m = _TAG_NAME_RE.match(tag)
    return f"<{m.group(1)}>" if m else tag


class QAIssue(NamedTuple):
    level: str   # "error" | "warning"
    code:  str   # short identifier, e.g. "UNTRANSLATED"
    msg:   str   # human-readable detail


# ── Stats ─────────────────────────────────────────────────────────────────────

class Stats:
    def __init__(self) -> None:
        self.files_ok:             int       = 0
        self.files_fail:           int       = 0
        self.segments_created:     int       = 0
        self.skipped_dup:          int       = 0
        self.skipped_null:         int       = 0
        self.errors:               list[str] = []
        self.warnings:             list[str] = []
        # QA
        self.qa_flagged:           int       = 0   # segments with ≥1 QA issue
        self.qa_excluded:          int       = 0   # removed by --strict

    # ---- helpers ----

    def err(self, msg: str)  -> None: self.errors.append(msg)
    def warn(self, msg: str) -> None: self.warnings.append(msg)

    # ---- report ----

    def report(self) -> None:
        W   = 60
        SEP = "═" * W
        THN = "─" * W

        def row(label: str, val: Any) -> str:
            return f"  {label:<30} {val}"

        ok_sym = "✓" if not self.files_fail else "!"
        lines = [
            "",
            SEP,
            f"  {'PROCESSING SUMMARY':^{W - 4}}",
            SEP,
            row("Files processed:",        self.files_ok + self.files_fail),
            row(f"  {ok_sym}  loaded OK:", self.files_ok),
            row(  "  ✗  load failures:",   self.files_fail),
            THN,
            row("Segments created:",       self.segments_created),
            row("Skipped — duplicate:",    self.skipped_dup),
            row("Skipped — null/empty:",   self.skipped_null),
            THN,
            row("QA issues found:",        self.qa_flagged),
            row("  excluded (--strict):",  self.qa_excluded),
            THN,
            row("Struct errors:",          len(self.errors)),
            row("Struct warnings:",        len(self.warnings)),
        ]

        def _list(header: str, items: list[str]) -> None:
            if not items:
                return
            lines.append(THN)
            lines.append(f"  {header}")
            for msg in items[:30]:
                lines.append(f"    • {msg}")
            if len(items) > 30:
                lines.append(f"    … and {len(items) - 30} more")

        _list("Structural errors:",   self.errors)
        _list("Structural warnings:", self.warnings)
        lines.append(SEP)
        print("\n".join(lines))


# ── JSON loading ──────────────────────────────────────────────────────────────

def _repair_json(text: str) -> str:
    """
    Fix unescaped double-quote characters inside JSON string values.

    Some source files use ``""text""`` as a typographic quoting convention.
    This scanner identifies which quotes are structural (open/close a JSON
    string) vs. typographic (embedded in text), and escapes the latter as \\".

    A ``"`` is treated as structural (end-of-string) when the next
    non-whitespace character after it is a JSON delimiter: ``,``, ``}``,
    ``]``, ``:``, or end of input.  Any other ``"`` found inside an open
    string is written as ``\\"``.
    """
    out: list[str] = []
    i = 0
    n = len(text)
    in_string = False

    while i < n:
        c = text[i]

        if not in_string:
            out.append(c)
            if c == '"':
                in_string = True
            i += 1
            continue

        # ── Inside a string ──────────────────────────────────────────────
        if c == '\\':
            # Escaped sequence — copy both characters verbatim
            out.append(c)
            i += 1
            if i < n:
                out.append(text[i])
                i += 1
            continue

        if c == '"':
            # Peek at the next non-whitespace character to decide:
            # structural closing quote, or unescaped typographic quote?
            j = i + 1
            while j < n and text[j] in ' \t\r\n':
                j += 1
            next_ch = text[j] if j < n else ''
            if next_ch in (',', '}', ']', ':') or j >= n:
                out.append('"')        # real end of string
                in_string = False
            else:
                out.append('\\"')      # unescaped internal quote — escape it
            i += 1
            continue

        out.append(c)
        i += 1

    return ''.join(out)


def load_json(path: Path, label: str, stats: Stats) -> Any | None:
    """Load a JSON file; return None and record stats on failure."""
    try:
        raw = path.read_bytes()
    except OSError as exc:
        stats.files_fail += 1
        stats.err(f"[{label}] Cannot read '{path}': {exc}")
        return None

    try:
        text = raw.decode("utf-8")
    except UnicodeDecodeError as exc:
        stats.files_fail += 1
        stats.err(f"[{label}] Encoding error in '{path.name}': {exc}")
        return None

    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        # First attempt failed — try auto-repairing unescaped quotes
        try:
            data = json.loads(_repair_json(text))
            stats.warn(f"[{label}] Auto-repaired unescaped quotes in '{path.name}'")
        except json.JSONDecodeError as exc2:
            stats.files_fail += 1
            stats.err(
                f"[{label}] JSON syntax error in '{path.name}' "
                f"at line {exc2.lineno}, col {exc2.colno}: {exc2.msg}"
            )
            return None

    stats.files_ok += 1
    print(f"  ✓  [{label}] {path.name}  ({len(raw):,} bytes)")
    return data


# ── Lint ──────────────────────────────────────────────────────────────────────

def lint(node: Any, path: str, stats: Stats) -> None:
    """
    Recursively validate a JSON node.

    Checks
    ------
    • UUID fields match the standard UUID format.
    • Language-tagged fields (ending in _en / _de / _ru / En / De / Ru)
      must be str, null, or absent — never int / float / list / dict.
    • String values in language fields must be non-empty if not null
      (empty string is treated as a warning, not an error).
    """
    if isinstance(node, dict):
        for key, val in node.items():
            cpath = f"{path}.{key}" if path else key

            # UUID format
            if UUID_KEY_RE.search(key):
                if val is not None and not UUID_RE.match(str(val)):
                    stats.err(f"Bad UUID at {cpath!r}: {val!r}")

            # Language field type check
            is_lang = any(
                key.endswith(s) or key.endswith(c)
                for s, c, _ in LANGS.values()
            )
            if is_lang:
                if val is not None and not isinstance(val, str):
                    stats.err(
                        f"Non-string value in lang field {cpath!r}: "
                        f"{type(val).__name__} = {val!r}"
                    )
                elif isinstance(val, str) and val.strip() == "":
                    stats.warn(f"Empty string in lang field {cpath!r}")

            lint(val, cpath, stats)

    elif isinstance(node, list):
        for i, item in enumerate(node):
            lint(item, f"{path}[{i}]", stats)


# ── Segment data class ────────────────────────────────────────────────────────

class Segment:
    """A matched EN → DE / RU translation unit."""
    __slots__ = ("path", "ctx_uuid", "en", "de", "ru", "qa_de", "qa_ru")

    def __init__(
        self,
        path:     str,
        ctx_uuid: str | None,
        en:       str,
        de:       str | None,
        ru:       str | None,
    ) -> None:
        self.path     = path
        self.ctx_uuid = ctx_uuid
        self.en       = en
        self.de       = de
        self.ru       = ru
        self.qa_de: list[QAIssue] = []
        self.qa_ru: list[QAIssue] = []


# ── Parallel tree walk ────────────────────────────────────────────────────────

def _nearest_uuid(node: dict) -> str | None:
    for key, val in node.items():
        if UUID_KEY_RE.search(key) and val and UUID_RE.match(str(val)):
            return str(val)
    return None


def _is_en_key(key: str) -> bool:
    s, c, _ = LANGS["en"]
    return key.endswith(s) or key.endswith(c)

def _is_de_key(key: str) -> bool:
    s, c, _ = LANGS["de"]
    return key.endswith(s) or key.endswith(c)

def _is_ru_key(key: str) -> bool:
    s, c, _ = LANGS["ru"]
    return key.endswith(s) or key.endswith(c)


def _peer_key(en_key: str, tgt_lang: str) -> str:
    """Derive the DE or RU peer key from an EN key."""
    s_en, c_en, _ = LANGS["en"]
    s_tgt, c_tgt, _ = LANGS[tgt_lang]
    if en_key.endswith(s_en):
        return en_key[: -len(s_en)] + s_tgt
    return en_key[: -len(c_en)] + c_tgt   # camelCase


def _walk(
    en_node:     Any,
    de_node:     Any,
    ru_node:     Any,
    path:        str,
    ctx_uuid:    str | None,
    same_keys:   bool,
    stats:       Stats,
    pbar,
    out:         list[Segment],
    text_fields:  set[str]       = frozenset(),
    active_langs: frozenset[str] = frozenset({"de", "ru"}),
) -> None:
    """
    Recursively walk all three trees in lock-step.

    For every *_en / *En key in en_node:
      • resolve the DE target from de_node (key *_de / *De, or same key if --same-keys)
      • resolve the RU target from ru_node (key *_ru / *Ru, or same key if --same-keys)
      • emit a Segment

    All other keys are recursed into (lists and dicts).
    *_de / *Ru keys are skipped during recursion (they're target values, not structure).
    """
    if isinstance(en_node, dict):
        # Narrow-scope UUID for this level
        uuid = _nearest_uuid(en_node) or ctx_uuid

        de_d = de_node if isinstance(de_node, dict) else {}
        ru_d = ru_node if isinstance(ru_node, dict) else {}

        # Structure consistency check (key-sets of non-language keys must match).
        # Strip ALL language-tagged keys before comparing: each file legitimately
        # holds only its own language variants (*_en / *_de / *_ru / camelCase
        # equivalents), so those must never be treated as structural mismatches.
        def _structural_keys(d: dict) -> set[str]:
            return {
                k for k in d
                if not _is_en_key(k) and not _is_de_key(k) and not _is_ru_key(k)
            }

        en_struct = _structural_keys(en_node)
        if "de" in active_langs:
            if _structural_keys(de_d) != en_struct:
                missing = en_struct - _structural_keys(de_d)
                extra   = _structural_keys(de_d) - en_struct
                if missing: stats.warn(f"DE missing keys at {path!r}: {sorted(missing)}")
                if extra:   stats.warn(f"DE extra keys at {path!r}: {sorted(extra)}")
        if "ru" in active_langs:
            if _structural_keys(ru_d) != en_struct:
                missing = en_struct - _structural_keys(ru_d)
                extra   = _structural_keys(ru_d) - en_struct
                if missing: stats.warn(f"RU missing keys at {path!r}: {sorted(missing)}")
                if extra:   stats.warn(f"RU extra keys at {path!r}: {sorted(extra)}")

        for key, en_val in en_node.items():
            cpath = f"{path}.{key}" if path else key

            if _is_en_key(key):
                # ── Extract translation pair ──────────────────────────────
                if same_keys:
                    # All three files use the same field name; values differ
                    de_val = de_d.get(key)
                    ru_val = ru_d.get(key)
                else:
                    de_key = _peer_key(key, "de")
                    ru_key = _peer_key(key, "ru")
                    de_val = de_d.get(de_key)
                    ru_val = ru_d.get(ru_key)
                    # Fallback: same key name (covers partial-migration files)
                    if de_val is None and de_key not in de_d:
                        de_val = de_d.get(key)
                    if ru_val is None and ru_key not in ru_d:
                        ru_val = ru_d.get(key)

                en_text = _clean_text(en_val) if isinstance(en_val, str) and en_val else None

                if not en_text:
                    stats.skipped_null += 1
                    pbar.update(1)
                    continue

                de_text = _clean_text(de_val) if isinstance(de_val, str) and de_val else None
                ru_text = _clean_text(ru_val) if isinstance(ru_val, str) and ru_val else None

                if "de" in active_langs and de_text is None:
                    stats.warn(f"No DE translation for {cpath!r}")
                if "ru" in active_langs and ru_text is None:
                    stats.warn(f"No RU translation for {cpath!r}")

                out.append(Segment(cpath, uuid, en_text, de_text, ru_text))
                pbar.update(1)

            elif key in text_fields:
                # ── Plain-name translatable field (no language suffix) ─────
                # Value is read from each language file at the same JSON path.
                # The UUID in the enclosing object confirms alignment.
                en_text = _clean_text(en_val) if isinstance(en_val, str) and en_val else None

                if not en_text:
                    stats.skipped_null += 1
                    pbar.update(1)
                    continue

                de_raw  = de_d.get(key)
                ru_raw  = ru_d.get(key)
                de_text = _clean_text(de_raw) if isinstance(de_raw, str) and de_raw else None
                ru_text = _clean_text(ru_raw) if isinstance(ru_raw, str) and ru_raw else None

                if "de" in active_langs and de_text is None:
                    stats.warn(f"No DE translation for text-field {cpath!r}")
                if "ru" in active_langs and ru_text is None:
                    stats.warn(f"No RU translation for text-field {cpath!r}")

                out.append(Segment(cpath, uuid, en_text, de_text, ru_text))
                pbar.update(1)

            elif not _is_de_key(key) and not _is_ru_key(key):
                # Recurse into structural (non-language) fields only
                _walk(
                    en_val,
                    de_d.get(key),
                    ru_d.get(key),
                    cpath,
                    uuid,
                    same_keys,
                    stats,
                    pbar,
                    out,
                    text_fields,
                    active_langs,
                )

    elif isinstance(en_node, list):
        de_l = de_node if isinstance(de_node, list) else []
        ru_l = ru_node if isinstance(ru_node, list) else []

        if "de" in active_langs and len(de_l) != len(en_node):
            stats.warn(
                f"Array length mismatch at {path!r}: EN={len(en_node)}, DE={len(de_l)}"
            )
        if "ru" in active_langs and len(ru_l) != len(en_node):
            stats.warn(
                f"Array length mismatch at {path!r}: EN={len(en_node)}, RU={len(ru_l)}"
            )

        for i, en_item in enumerate(en_node):
            _walk(
                en_item,
                de_l[i] if i < len(de_l) else None,
                ru_l[i] if i < len(ru_l) else None,
                f"{path}[{i}]",
                ctx_uuid,
                same_keys,
                stats,
                pbar,
                out,
                text_fields,
                active_langs,
            )


# ── Pre-count (for a meaningful progress bar) ─────────────────────────────────

def _count_en_fields(node: Any, text_fields: set[str] = frozenset()) -> int:
    if isinstance(node, dict):
        total = 0
        for key, val in node.items():
            if _is_en_key(key):
                total += 1
            elif key in text_fields and isinstance(val, str) and val.strip():
                total += 1
            elif not _is_de_key(key) and not _is_ru_key(key):
                total += _count_en_fields(val, text_fields)
        return total
    if isinstance(node, list):
        return sum(_count_en_fields(item, text_fields) for item in node)
    return 0


# ── Deduplication ─────────────────────────────────────────────────────────────

def deduplicate(raw: list[Segment], stats: Stats) -> list[Segment]:
    """
    Drop segments whose (en, de, ru) triplet was already seen.
    First occurrence wins; all later ones are counted as duplicates.
    """
    seen: set[str] = set()
    out:  list[Segment] = []
    for seg in raw:
        key = hashlib.md5(
            f"{seg.en}\x00{seg.de or ''}\x00{seg.ru or ''}".encode(),
            usedforsecurity=False,
        ).hexdigest()
        if key in seen:
            stats.skipped_dup += 1
        else:
            seen.add(key)
            out.append(seg)
    return out


# ── Segment ID ────────────────────────────────────────────────────────────────

def seg_id(en_text: str) -> str:
    return hashlib.md5(en_text.encode(), usedforsecurity=False).hexdigest()[:12]


# ── QA checks ─────────────────────────────────────────────────────────────────

def qa_check_pair(en: str, tgt: str, tgt_lang: str, path: str) -> list[QAIssue]:
    """
    Run all QA checks on one EN→target pair.  Returns a (possibly empty)
    list of QAIssue named tuples.

    Checks
    ------
    UNTRANSLATED      (warning)  Target text is identical to source.
    LENGTH_RATIO      (warning)  Character-count ratio is outside expected bounds.
    NUMBER_MISMATCH   (error)    Digit sequences present in EN are absent/changed
                                 in the target, or extra digits appear.
    PLACEHOLDER_MISMATCH (error) Interpolation markers ({var}, %s, …) differ.
    NO_CYRILLIC       (warning)  RU translation contains no Cyrillic characters.
    WRONG_SCRIPT      (warning)  DE translation contains Cyrillic characters
                                 (likely a swapped or mislabelled file).
    TAG_MISMATCH      (error)    HTML/XML tags present in EN differ in target.
    """
    issues: list[QAIssue] = []

    # 1. Untranslated — target identical to source
    #    Skip short strings and pure-uppercase abbreviations (OK, DNA, MRI …)
    if (
        len(en) > 5
        and _LATIN_LC_RE.search(en)
        and en.strip().lower() == tgt.strip().lower()
    ):
        issues.append(QAIssue(
            "warning", "UNTRANSLATED",
            f"{path}: target is identical to source — {en!r:.80}",
        ))

    # 2. Length ratio
    #    Generous bounds: EN→DE German runs ~30 % longer; EN→RU varies widely.
    #    Only apply to strings of 10+ chars to avoid false positives on short labels.
    if len(en) >= 10:
        ratio = len(tgt) / len(en)
        lo, hi = (0.25, 4.0) if tgt_lang == "de" else (0.20, 5.0)
        if not (lo <= ratio <= hi):
            issues.append(QAIssue(
                "warning", "LENGTH_RATIO",
                f"{path}: {tgt_lang.upper()}/EN length ratio {ratio:.2f} "
                f"(EN={len(en)} chars, {tgt_lang.upper()}={len(tgt)} chars)",
            ))

    # 3. Number preservation
    #    Compare multisets of digit runs so "1 234" and "1234" both surface.
    en_nums  = Counter(_NUM_RE.findall(en))
    tgt_nums = Counter(_NUM_RE.findall(tgt))
    missing  = en_nums - tgt_nums
    extra    = tgt_nums - en_nums
    if missing or extra:
        parts = []
        if missing: parts.append(f"missing {dict(missing)}")
        if extra:   parts.append(f"extra {dict(extra)}")
        issues.append(QAIssue(
            "error", "NUMBER_MISMATCH",
            f"{path}: numbers differ in {tgt_lang.upper()} — {', '.join(parts)}",
        ))

    # 4. Placeholder preservation
    en_ph  = Counter(_PH_RE.findall(en))
    tgt_ph = Counter(_PH_RE.findall(tgt))
    if en_ph != tgt_ph:
        issues.append(QAIssue(
            "error", "PLACEHOLDER_MISMATCH",
            f"{path}: placeholders differ — "
            f"EN {dict(en_ph)}, {tgt_lang.upper()} {dict(tgt_ph)}",
        ))

    # 5. Cyrillic script check for Russian
    #    Only flag real prose (source has lowercase Latin); skip abbreviations.
    if tgt_lang == "ru" and len(en) > 3 and _LATIN_LC_RE.search(en):
        if not _CYRILLIC_RE.search(tgt):
            issues.append(QAIssue(
                "warning", "NO_CYRILLIC",
                f"{path}: RU target has no Cyrillic — {tgt!r:.80}",
            ))

    # 6. Wrong script in DE (Cyrillic in a German field → likely swapped file)
    if tgt_lang == "de" and _CYRILLIC_RE.search(tgt):
        issues.append(QAIssue(
            "warning", "WRONG_SCRIPT",
            f"{path}: DE target contains Cyrillic characters — {tgt!r:.80}",
        ))

    # 7. HTML / XML tag preservation
    # Tags are normalised to names only (<d=Holotopy> → <d>) so that
    # translated section labels don't generate spurious mismatches.
    en_tags  = Counter(_strip_tag_attrs(t) for t in _TAG_RE.findall(en))
    tgt_tags = Counter(_strip_tag_attrs(t) for t in _TAG_RE.findall(tgt))
    if en_tags != tgt_tags:
        issues.append(QAIssue(
            "error", "TAG_MISMATCH",
            f"{path}: HTML tags differ — EN {dict(en_tags)}, "
            f"{tgt_lang.upper()} {dict(tgt_tags)}",
        ))

    return issues


def run_qa(segments: list[Segment], strict: bool, stats: Stats) -> list[Segment]:
    """
    Run QA checks on every segment.  Fills seg.qa_de and seg.qa_ru in-place.

    strict=False  All segments are passed through; issues are counted and
                  recorded in stats but do not affect the output list.
    strict=True   Segments that have at least one QA issue (any level) are
                  excluded from the returned list and counted in stats.qa_excluded.
    """
    out: list[Segment] = []

    for seg in segments:
        if seg.de:
            seg.qa_de = qa_check_pair(seg.en, seg.de, "de", seg.path)
        if seg.ru:
            seg.qa_ru = qa_check_pair(seg.en, seg.ru, "ru", seg.path)

        all_issues = seg.qa_de + seg.qa_ru

        if all_issues:
            stats.qa_flagged += 1

        if strict and all_issues:
            stats.qa_excluded += 1
        else:
            out.append(seg)

    return out


# ── TMX writer ────────────────────────────────────────────────────────────────

def write_tmx(
    segments:  list[Segment],
    src_bcp47: str,
    tgt_bcp47: str,
    tgt_attr:  str,       # "de" or "ru"
    out_path:  Path,
    pbar,
) -> int:
    root = ET.Element("tmx", version="1.4")
    ET.SubElement(
        root, "header",
        creationtool=TOOL_NAME,
        creationtoolversion=TOOL_VERSION,
        datatype="plaintext",
        segtype="sentence",
        adminlang="en",
        srclang=src_bcp47,
        **{"o-tmf": TOOL_NAME},
    )
    body = ET.SubElement(root, "body")
    written = 0

    for seg in segments:
        tgt_text: str | None = getattr(seg, tgt_attr)
        if not tgt_text:
            pbar.update(1)
            continue

        tu = ET.SubElement(body, "tu")

        src_tuv = ET.SubElement(tu, "tuv", **{"xml:lang": src_bcp47})
        ET.SubElement(src_tuv, "seg").text = seg.en

        tgt_tuv = ET.SubElement(tu, "tuv", **{"xml:lang": tgt_bcp47})
        ET.SubElement(tgt_tuv, "seg").text = tgt_text

        written += 1
        pbar.update(1)

    # Pretty-print with DOCTYPE
    raw_xml = ET.tostring(root, encoding="unicode")
    dom     = minidom.parseString(raw_xml)
    pretty  = dom.toprettyxml(indent="  ", encoding="UTF-8").decode("UTF-8")
    pretty  = pretty.replace(
        '<?xml version="1.0" ?>\n',
        '<?xml version="1.0" encoding="UTF-8"?>\n'
        '<!DOCTYPE tmx SYSTEM "tmx14.dtd">\n',
    )
    out_path.write_text(pretty, encoding="utf-8")
    return written


# ── XLSX writer ───────────────────────────────────────────────────────────────

_HDR_FILL  = None
_HDR_FONT  = None
_WRAP      = None
_ERR_FILL  = None   # missing translation / QA error
_WARN_FILL = None   # QA warning

if HAS_OPENPYXL:
    _HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
    _HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
    _WRAP      = Alignment(wrap_text=True, vertical="top")
    _ERR_FILL  = PatternFill("solid", fgColor="FFD0D0")
    _WARN_FILL = PatternFill("solid", fgColor="FFF3CD")


def _apply_header(ws, columns: list[str]) -> None:
    ws.append(columns)
    for cell in ws[1]:
        cell.font      = _HDR_FONT
        cell.fill      = _HDR_FILL
        cell.alignment = _WRAP
    ws.row_dimensions[1].height = 22


def write_xlsx(
    segments: list[Segment],
    out_path: Path,
    pbar,
    has_de:   bool = True,
    has_ru:   bool = True,
) -> int:
    if not HAS_OPENPYXL:
        print("  ⚠  openpyxl not installed — skipping XLSX output", file=sys.stderr)
        return 0

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    col_widths = [6, 14, 50, 50, 36, 55, 40]
    headers    = ["#", "Segment ID", "EN (source)", "Target",
                  "Context UUID", "JSON path", "QA issues"]

    def _fill_sheet(ws, tgt_attr: str, qa_attr: str) -> int:
        _apply_header(ws, headers)
        count = 0
        for seg in segments:
            tgt: str | None           = getattr(seg, tgt_attr)
            issues: list[QAIssue]     = getattr(seg, qa_attr)
            if not tgt:
                pbar.update(1)
                continue
            count += 1
            row_idx = count + 1

            qa_cell = "; ".join(f"[{i.level.upper()}] {i.code}" for i in issues)
            ws.append([count, seg_id(seg.en), seg.en, tgt,
                       seg.ctx_uuid or "", seg.path, qa_cell])

            for cell in ws[row_idx]:
                cell.alignment = _WRAP

            # Row highlight: error (red) > warning (amber) > clean (none)
            has_error   = any(i.level == "error"   for i in issues)
            has_warning = any(i.level == "warning"  for i in issues)
            fill = _ERR_FILL if has_error else (_WARN_FILL if has_warning else None)
            if fill:
                for cell in ws[row_idx]:
                    cell.fill = fill

            pbar.update(1)
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        return count

    n_de = n_ru = 0
    if has_de:
        ws_de = wb.create_sheet("EN-DE")
        n_de  = _fill_sheet(ws_de, "de", "qa_de")
    if has_ru:
        ws_ru = wb.create_sheet("EN-RU")
        n_ru  = _fill_sheet(ws_ru, "ru", "qa_ru")

    # Summary sheet (front)
    ws_s = wb.create_sheet("Summary", 0)
    _apply_header(ws_s, ["Metric", "Value"])
    summary_rows: list[tuple] = [("Total unique segments", len(segments))]
    if has_de: summary_rows.append(("EN→DE pairs", n_de))
    if has_ru: summary_rows.append(("EN→RU pairs", n_ru))
    summary_rows.append(("Generated at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    summary_rows.append(("Tool",          f"{TOOL_NAME} {TOOL_VERSION}"))
    for row in summary_rows:
        ws_s.append(row)
    ws_s.column_dimensions["A"].width = 24
    ws_s.column_dimensions["B"].width = 26
    ws_s.freeze_panes = "A2"

    wb.save(out_path)
    return len(segments)


# ── Triplet resolver ──────────────────────────────────────────────────────────

def resolve_triplets(
    en_path: Path, de_path: Path | None, ru_path: Path | None
) -> list[tuple[Path, Path | None, Path | None]]:
    """
    Return a list of (en_file, de_file, ru_file) tuples.

    de_path / ru_path may be None when a language was not supplied.

    File mode  : en_path is a file → single-element list.
    Dir mode   : en_path is a directory → recursively glob *.json, then
                 resolve each file's relative path inside de_path / ru_path.
    Mixed paths (file + dir) are rejected early with a clear error.
    """
    en_is_dir = en_path.is_dir()

    # Each provided target must be the same kind (file vs dir) as EN.
    for tgt_path, flag in ((de_path, "--de"), (ru_path, "--ru")):
        if tgt_path is None:
            continue
        if tgt_path.is_dir() != en_is_dir:
            sys.exit(
                f"Error: --en and {flag} must both be files or both be directories.\n"
                f"  --en  {'dir' if en_is_dir else 'file'} ({en_path})\n"
                f"  {flag}  {'dir' if tgt_path.is_dir() else 'file'} ({tgt_path})"
            )

    if not en_is_dir:
        # ── file mode ────────────────────────────────────────────────────────
        if not en_path.exists():
            sys.exit(f"Error: --en file not found: {en_path}")
        if not en_path.is_file():
            sys.exit(f"Error: --en path is not a file: {en_path}")
        for tgt_path, flag in ((de_path, "--de"), (ru_path, "--ru")):
            if tgt_path is not None:
                if not tgt_path.exists():
                    sys.exit(f"Error: {flag} file not found: {tgt_path}")
                if not tgt_path.is_file():
                    sys.exit(f"Error: {flag} path is not a file: {tgt_path}")
        return [(en_path, de_path, ru_path)]

    # ── directory mode ───────────────────────────────────────────────────────
    en_files = sorted(en_path.rglob("*.json"))
    if not en_files:
        sys.exit(f"Error: no *.json files found under {en_path}")

    triplets: list[tuple[Path, Path | None, Path | None]] = []
    for en_file in en_files:
        rel     = en_file.relative_to(en_path)
        de_file = de_path / rel if de_path is not None else None
        ru_file = ru_path / rel if ru_path is not None else None
        triplets.append((en_file, de_file, ru_file))

    return triplets


# ── CLI ───────────────────────────────────────────────────────────────────────

def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="json2tm",
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument("--en",        required=True, metavar="FILE/DIR",
                   help="English source: a single JSON file or a directory of JSON files")
    p.add_argument("--de",        default=None, metavar="FILE/DIR",
                   help="German target: a single JSON file or a directory (must match --en). "
                        "Optional — omit to produce EN→RU output only.")
    p.add_argument("--ru",        default=None, metavar="FILE/DIR",
                   help="Russian target: a single JSON file or a directory (must match --en). "
                        "Optional — omit to produce EN→DE output only.")
    p.add_argument("--out",       default="output", metavar="DIR",
                   help="Output directory  (default: output/)")
    p.add_argument("--same-keys", action="store_true",
                   help="All three files use identical field names; "
                        "the text values differ per language.")
    p.add_argument("--strict",    action="store_true",
                   help="Exclude segments that fail any QA check from all output. "
                        "Without this flag QA issues are reported but segments are kept.")
    p.add_argument("--text-fields", default="", metavar="FIELD[,FIELD,...]",
                   help="Comma-separated list of plain (non-language-suffixed) field names "
                        "whose values are translatable text.  The same field name must exist "
                        "in all three language files; each file holds the text in its own "
                        "language.  Example: --text-fields categoryName,pathology_name")
    p.add_argument("--no-tmx",    action="store_true",
                   help="Skip TMX output")
    p.add_argument("--no-xlsx",   action="store_true",
                   help="Skip XLSX output")
    return p


def main() -> None:
    args    = build_parser().parse_args()
    stats   = Stats()
    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)

    text_fields: set[str] = (
        {f.strip() for f in args.text_fields.split(",") if f.strip()}
        if args.text_fields else set()
    )

    if args.de is None and args.ru is None:
        build_parser().error("at least one of --de or --ru is required")

    has_de = args.de is not None
    has_ru = args.ru is not None
    active_langs: frozenset[str] = frozenset(
        lang for lang, present in (("de", has_de), ("ru", has_ru)) if present
    )

    # ── 1. Resolve file triplets ──────────────────────────────────────────────
    triplets = resolve_triplets(
        Path(args.en),
        Path(args.de) if args.de else None,
        Path(args.ru) if args.ru else None,
    )
    dir_mode = Path(args.en).is_dir()
    print(
        f"\n── {'Directory' if dir_mode else 'File'} mode"
        f"  ({len(triplets):,} file triplet{'s' if len(triplets) != 1 else ''}) "
        + "─" * 20
    )

    # ── 2. Load → Lint → Extract  (one triplet at a time) ────────────────────
    raw: list[Segment] = []

    for i, (en_p, de_p, ru_p) in enumerate(triplets, 1):
        prefix = f"[{i}/{len(triplets)}]" if len(triplets) > 1 else ""

        # Load
        en_data = load_json(en_p, f"{prefix} EN", stats)
        de_data = load_json(de_p, f"{prefix} DE", stats) if de_p is not None else None
        ru_data = load_json(ru_p, f"{prefix} RU", stats) if ru_p is not None else None

        if (en_data is None
                or (has_de and de_data is None)
                or (has_ru and ru_data is None)):
            print(f"  ⚠  Skipping triplet {i} (load failure)")
            continue

        # Lint
        _lint_pairs = [(en_data, "EN")]
        if has_de: _lint_pairs.append((de_data, "DE"))
        if has_ru: _lint_pairs.append((ru_data, "RU"))
        for data, label in _lint_pairs:
            before = len(stats.errors)
            lint(data, label, stats)
            n_new  = len(stats.errors) - before
            sym    = "✓" if n_new == 0 else f"✗ ({n_new} errors)"
            print(f"  {prefix} lint [{label}]  {sym}")

        # Extract
        total_en = _count_en_fields(en_data, text_fields)
        with _tqdm_cls(
            total=total_en,
            desc=f"  {prefix} extracting",
            unit=" fields",
            dynamic_ncols=True,
            colour="cyan",
            disable=not HAS_TQDM,
        ) as pbar:
            _walk(en_data, de_data, ru_data, "", None, args.same_keys, stats, pbar, raw,
                  text_fields, active_langs)

    if not raw:
        print("\nNo segments extracted. Aborting.")
        stats.report()
        sys.exit(1)

    print(f"\n  {len(raw):,} candidate segments collected across all files")

    # ── 5. Deduplicate ────────────────────────────────────────────────────────
    print("\n── Deduplicating ────────────────────────────────────────────")
    segments = deduplicate(raw, stats)
    print(f"  {stats.skipped_dup:,} duplicates removed → {len(segments):,} unique segments")

    # ── 6. QA ─────────────────────────────────────────────────────────────────
    print("\n── QA checks ────────────────────────────────────────────────")
    segments = run_qa(segments, strict=args.strict, stats=stats)
    stats.segments_created = len(segments)
    qa_sym = "✓" if stats.qa_flagged == 0 else f"⚠  {stats.qa_flagged:,} issue(s) found"
    print(f"  {qa_sym}")
    if stats.qa_excluded:
        print(f"  {stats.qa_excluded:,} segments excluded by --strict")
    print(f"  {len(segments):,} segments ready for output")

    # ── 8. Write TMX ──────────────────────────────────────────────────────────
    if not args.no_tmx:
        print("\n── Writing TMX ──────────────────────────────────────────────")
        pairs = []
        if has_de: pairs.append(("de", "de", "en-de.tmx"))
        if has_ru: pairs.append(("ru", "ru", "en-ru.tmx"))
        for lang, attr, fname in pairs:
            out_path = out_dir / fname
            with _tqdm_cls(
                total=len(segments),
                desc=f"  {fname}",
                unit=" TUs",
                dynamic_ncols=True,
                colour="green",
                disable=not HAS_TQDM,
            ) as pbar:
                n = write_tmx(segments, "en", lang, attr, out_path, pbar)
            size_kb = out_path.stat().st_size / 1024
            print(f"  ✓  {fname}  ({n:,} TUs, {size_kb:.1f} KB)")

    # ── 9. Write XLSX ─────────────────────────────────────────────────────────
    if not args.no_xlsx:
        print("\n── Writing XLSX ─────────────────────────────────────────────")
        xlsx_path = out_dir / "translations.xlsx"
        # 2 sheets × segments each
        with _tqdm_cls(
            total=len(segments) * (int(has_de) + int(has_ru)),
            desc="  translations.xlsx",
            unit=" rows",
            dynamic_ncols=True,
            colour="yellow",
            disable=not HAS_TQDM,
        ) as pbar:
            n = write_xlsx(segments, xlsx_path, pbar, has_de=has_de, has_ru=has_ru)
        if xlsx_path.exists():
            size_kb = xlsx_path.stat().st_size / 1024
            print(f"  ✓  translations.xlsx  ({n:,} segments, {size_kb:.1f} KB)")

    # ── 10. Summary ───────────────────────────────────────────────────────────
    stats.report()
    sys.exit(1 if stats.errors else 0)


if __name__ == "__main__":
    main()
